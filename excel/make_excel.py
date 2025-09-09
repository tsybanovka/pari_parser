import pandas as pd
from openpyxl import load_workbook

def write_data(filepath:str, data_in:list):
    # Создаем словарь с заголовками столбцов
    data_out = {
        'command1': [],
        'command2': [],
        "score(1:2)": [],
        's1(1:2)': [],
        's2(1:2)': [],
        's3(1:2)': [],
        's4(1:2)': [],
        's5(1:2)': []
    }

    for (command1, command2), score, dt in data_in:
        data_out["command1"].append(command1)
        data_out["command2"].append(command2)

        data_out["score(1:2)"].append(score)

        for i in range(1, len(dt)+1):
            data_out[f"s{i}(1:2)"].append(dt[i-1])

    # Преобразуем словарь в DataFrame
    df = pd.DataFrame(data_out)

    df.to_excel(filepath, index=False)

def append_data(filepath:str, data_in:list):
    from openpyxl import load_workbook

    # Открываем существующий файл
    wb = load_workbook(filename=filepath)
    ws = wb.active  # Активируем первый доступный лист

    for (command1, command2), score, dt in data_in:
        ws.append([command1, command2, score, *dt])

    # Сохраняем изменения
    wb.save(filepath)
